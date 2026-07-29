#!/usr/bin/env python3
"""Regenerate spreadsheet-sample.xlsx, the fixture for the XLSX extraction tests.

    uvx --with openpyxl python test/fixtures/files/generate-spreadsheet-sample.py

Python because openpyxl can write the awkward cells this grid needs — a
formula-error cell in particular — and because a devDependency whose only job is
building a 5 KB fixture is not worth carrying in the JS tree.

The grid is deliberately awkward. Every column carries a cell whose meaning
lives in its *number format* rather than its value, which is the class of thing
the extraction path has to preserve; the single round-trip assertion in
test/unit/files/extract.test.ts names the exact text produced here:

  Note     a comma, embedded quotes, a newline, non-ASCII, and a padded string
  Opened   date-only, stored as a whole-day serial
  Checked  datetime
  Shift    time-only, a serial below 1
  Ratio    a percent-formatted number, and one formula-error cell (#N/A)
"""

import datetime
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook

OUT = Path(__file__).with_name("spreadsheet-sample.xlsx")
ERROR_CELL = "H3"

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

ws.append(["Region", "Units", "Note", "Opened", "Checked", "Shift", "Active", "Ratio"])
rows = [
    ["West", 12, "red, small", datetime.datetime(2026, 1, 15),
     datetime.datetime(2026, 1, 15, 14, 30), datetime.time(14, 30), True, 0.5],
    # The Ratio placeholder here only exists so openpyxl emits the cell; it is
    # rewritten into a t="e" formula-error cell below.
    ["East", 7, 'has "quotes"', datetime.datetime(2026, 6, 1),
     datetime.datetime(2026, 6, 1, 9, 5), datetime.time(9, 5), False, 0],
    ["Ünicode ✓", 0, None, datetime.datetime(2025, 12, 31), None, None, True, None],
    ["Multi", 3.5, "line\nbreak", datetime.datetime(2026, 7, 4), None, None, False, None],
    ["Padded", 1, "  padded  ", None, None, None, True, None],
]
for row in rows:
    ws.append(row)

for r in range(2, 2 + len(rows)):
    ws[f"D{r}"].number_format = "mm-dd-yy"      # built-in id 14, date only
    ws[f"E{r}"].number_format = "m/d/yy h:mm"   # built-in id 22, datetime
    ws[f"F{r}"].number_format = "h:mm"          # built-in id 20, time only
ws["H2"].number_format = "0.0%"

wb.save(OUT)

# openpyxl cannot emit a t="e" cell, so the formula-error cell is patched into
# the sheet XML afterwards. Rewriting the archive entry-by-entry keeps every
# other part byte-identical to what openpyxl produced.
SHEET = "xl/worksheets/sheet1.xml"
with zipfile.ZipFile(OUT) as zf:
    entries = {n: zf.read(n) for n in zf.namelist()}
    order = zf.namelist()

xml = entries[SHEET].decode("utf-8")
start = xml.index(f'<c r="{ERROR_CELL}"')
end = xml.index("</c>", start) + len("</c>")
entries[SHEET] = (xml[:start] + f'<c r="{ERROR_CELL}" t="e"><v>#N/A</v></c>' + xml[end:]).encode("utf-8")

tmp = OUT.with_suffix(".tmp")
with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
    for name in order:
        zf.writestr(name, entries[name])
shutil.move(tmp, OUT)

print(f"wrote {OUT} ({OUT.stat().st_size} bytes)")
